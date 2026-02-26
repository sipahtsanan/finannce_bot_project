"""
จัดการไฟล์ Excel สำหรับบันทึกรายรับรายจ่าย
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


HEADER_FILL = PatternFill("solid", start_color="1F4E79")
SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
INCOME_FILL = PatternFill("solid", start_color="E2EFDA")
EXPENSE_FILL = PatternFill("solid", start_color="FCE4D6")
TOTAL_FILL = PatternFill("solid", start_color="FFF2CC")

WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BOLD = Font(name="Arial", bold=True, size=10)
NORMAL = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

THB = '#,##0.00'
THB_NEG = '#,##0.00;[Red]-#,##0.00'


def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


class ExcelManager:
    TRANSACTIONS_SHEET = "รายการ"
    MONTHLY_SHEET = "สรุปรายเดือน"
    YEARLY_SHEET = "สรุปรายปี"

    COLUMNS = ["วันที่", "หมวดหมู่", "รายละเอียด", "ประเภท", "จำนวนเงิน (บาท)"]

    def __init__(self, filepath: str):
        self.filepath = filepath
        if not os.path.exists(filepath):
            self._create_workbook()

    def _create_workbook(self):
        wb = Workbook()

        # ── Sheet 1: รายการ ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = self.TRANSACTIONS_SHEET

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "📊 บันทึกรายรับรายจ่าย"
        ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = HEADER_FILL
        ws["A1"].alignment = CENTER

        # Headers
        for col, header in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = WHITE_BOLD
            cell.fill = SUBHEADER_FILL
            cell.alignment = CENTER
            cell.border = border()

        # Column widths
        widths = [14, 22, 40, 12, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A3"
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

        # ── Sheet 2: สรุปรายเดือน ─────────────────────────────────────────────
        ws2 = wb.create_sheet(self.MONTHLY_SHEET)
        ws2.merge_cells("A1:D1")
        ws2["A1"] = "📅 สรุปรายจ่ายรายเดือน"
        ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws2["A1"].fill = HEADER_FILL
        ws2["A1"].alignment = CENTER
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 24
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 18
        ws2.row_dimensions[1].height = 28

        # ── Sheet 3: สรุปรายปี ────────────────────────────────────────────────
        ws3 = wb.create_sheet(self.YEARLY_SHEET)
        ws3.merge_cells("A1:D1")
        ws3["A1"] = "📆 สรุปค่าใช้จ่ายรายปี แยกตามหมวด"
        ws3["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws3["A1"].fill = HEADER_FILL
        ws3["A1"].alignment = CENTER
        ws3.column_dimensions["A"].width = 24
        ws3.column_dimensions["B"].width = 18
        ws3.column_dimensions["C"].width = 14
        ws3.column_dimensions["D"].width = 20
        ws3.row_dimensions[1].height = 28

        wb.save(self.filepath)

    def add_transaction(self, date: str, amount: float, category: str,
                        description: str, transaction_type: str = "expense"):
        wb = load_workbook(self.filepath)
        ws = wb[self.TRANSACTIONS_SHEET]

        # หาแถวถัดไป
        next_row = ws.max_row + 1
        if next_row < 3:
            next_row = 3

        # สีตาม type
        fill = INCOME_FILL if transaction_type == "income" else EXPENSE_FILL
        amount_val = amount if transaction_type == "income" else -amount

        values = [date, category, description, 
                  "รายได้" if transaction_type == "income" else "รายจ่าย",
                  amount_val]
        
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=next_row, column=col, value=val)
            cell.font = NORMAL
            cell.fill = fill
            cell.border = border()
            if col == 5:
                cell.number_format = THB_NEG
                cell.alignment = RIGHT
            elif col == 1:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

        wb.save(self.filepath)
        self._update_summary_sheets(wb)

    def _update_summary_sheets(self, wb: Workbook):
        ws = wb[self.TRANSACTIONS_SHEET]
        
        # อ่านข้อมูลทั้งหมด
        transactions = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:
                transactions.append({
                    "date": str(row[0])[:7],   # YYYY-MM
                    "year": str(row[0])[:4],   # YYYY
                    "category": row[1] or "อื่นๆ",
                    "type": row[3],
                    "amount": float(row[4] or 0)
                })

        # ── Monthly Summary ───────────────────────────────────────────────────
        ws2 = wb[self.MONTHLY_SHEET]
        ws2.delete_rows(2, ws2.max_row)

        # Headers
        headers = ["เดือน", "หมวดหมู่", "รายได้ (บาท)", "รายจ่าย (บาท)"]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(row=2, column=c, value=h)
            cell.font = WHITE_BOLD
            cell.fill = SUBHEADER_FILL
            cell.alignment = CENTER
            cell.border = border()

        # Group by month + category
        from collections import defaultdict
        monthly = defaultdict(lambda: defaultdict(lambda: {"income": 0, "expense": 0}))
        for t in transactions:
            key = t["date"]
            cat = t["category"]
            if t["amount"] > 0:
                monthly[key][cat]["income"] += t["amount"]
            else:
                monthly[key][cat]["expense"] += abs(t["amount"])

        row_num = 3
        for month in sorted(monthly.keys(), reverse=True):
            cats = monthly[month]
            month_total_inc = sum(v["income"] for v in cats.values())
            month_total_exp = sum(v["expense"] for v in cats.values())
            
            for cat, vals in sorted(cats.items()):
                for c, val in enumerate([month, cat, vals["income"], vals["expense"]], 1):
                    cell = ws2.cell(row=row_num, column=c, value=val)
                    cell.font = NORMAL
                    cell.border = border()
                    if c >= 3:
                        cell.number_format = THB
                        cell.alignment = RIGHT
                    else:
                        cell.alignment = CENTER if c == 1 else LEFT
                row_num += 1
            
            # แถว total ต่อเดือน
            total_cells = [ws2.cell(row=row_num, column=c) for c in range(1, 5)]
            total_cells[0].value = f"รวม {month}"
            total_cells[2].value = month_total_inc
            total_cells[3].value = month_total_exp
            for cell in total_cells:
                cell.font = BOLD
                cell.fill = TOTAL_FILL
                cell.border = border()
                if total_cells.index(cell) >= 2:
                    cell.number_format = THB
                    cell.alignment = RIGHT
                else:
                    cell.alignment = CENTER
            row_num += 1

        # ── Yearly Summary ────────────────────────────────────────────────────
        ws3 = wb[self.YEARLY_SHEET]
        ws3.delete_rows(2, ws3.max_row)

        headers3 = ["หมวดหมู่", "ยอดรายจ่าย (บาท)", "สัดส่วน (%)", "เฉลี่ย/เดือน (บาท)"]
        for c, h in enumerate(headers3, 1):
            cell = ws3.cell(row=2, column=c, value=h)
            cell.font = WHITE_BOLD
            cell.fill = SUBHEADER_FILL
            cell.alignment = CENTER
            cell.border = border()

        # Group by category (expense only)
        cat_totals = defaultdict(float)
        months_set = set()
        for t in transactions:
            if t["amount"] < 0:
                cat_totals[t["category"]] += abs(t["amount"])
                months_set.add(t["date"])

        total_expense = sum(cat_totals.values())
        num_months = max(len(months_set), 1)

        row_num3 = 3
        for cat, total in sorted(cat_totals.items(), key=lambda x: -x[1]):
            pct = (total / total_expense * 100) if total_expense else 0
            avg = total / num_months
            vals = [cat, total, pct, avg]
            for c, val in enumerate(vals, 1):
                cell = ws3.cell(row=row_num3, column=c, value=val)
                cell.font = NORMAL
                cell.border = border()
                if c == 1:
                    cell.alignment = LEFT
                elif c == 3:
                    cell.number_format = "0.0%"
                    cell.value = pct / 100
                    cell.alignment = CENTER
                else:
                    cell.number_format = THB
                    cell.alignment = RIGHT
            row_num3 += 1

        # แถว Grand Total
        if cat_totals:
            gt = ws3.cell(row=row_num3, column=1, value="💰 รวมทั้งหมด")
            gt.font = BOLD
            gt.fill = TOTAL_FILL
            gt.border = border()
            gt.alignment = LEFT
            
            gt2 = ws3.cell(row=row_num3, column=2, value=total_expense)
            gt2.font = BOLD
            gt2.fill = TOTAL_FILL
            gt2.border = border()
            gt2.number_format = THB
            gt2.alignment = RIGHT
            
            for c in [3, 4]:
                cell = ws3.cell(row=row_num3, column=c)
                cell.fill = TOTAL_FILL
                cell.border = border()

        wb.save(self.filepath)

    def get_monthly_summary(self) -> str:
        """คืนค่าสรุปเดือนนี้เป็น text"""
        wb = load_workbook(self.filepath, data_only=True)
        ws = wb[self.TRANSACTIONS_SHEET]
        
        current_month = datetime.now().strftime("%Y-%m")
        from collections import defaultdict
        cat_expense = defaultdict(float)
        cat_income = defaultdict(float)
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]:
                continue
            if str(row[0])[:7] == current_month:
                amount = float(row[4] or 0)
                cat = row[1] or "อื่นๆ"
                if amount > 0:
                    cat_income[cat] += amount
                else:
                    cat_expense[cat] += abs(amount)
        
        if not cat_expense and not cat_income:
            return f"📊 ยังไม่มีข้อมูลเดือน {current_month}"
        
        lines = [f"📊 *สรุปเดือน {current_month}*\n"]
        
        if cat_income:
            lines.append("💰 *รายได้:*")
            for cat, amt in sorted(cat_income.items(), key=lambda x: -x[1]):
                lines.append(f"  {cat}: {amt:,.0f} บาท")
            lines.append(f"  รวม: {sum(cat_income.values()):,.0f} บาท\n")
        
        if cat_expense:
            lines.append("💸 *รายจ่าย:*")
            for cat, amt in sorted(cat_expense.items(), key=lambda x: -x[1]):
                lines.append(f"  {cat}: {amt:,.0f} บาท")
            lines.append(f"  รวม: {sum(cat_expense.values()):,.0f} บาท")
        
        net = sum(cat_income.values()) - sum(cat_expense.values())
        lines.append(f"\n{'✅' if net >= 0 else '⚠️'} *คงเหลือ: {net:,.0f} บาท*")
        
        return "\n".join(lines)

    def get_yearly_summary(self) -> str:
        """คืนค่าสรุปรายปีเป็น text"""
        wb = load_workbook(self.filepath, data_only=True)
        ws = wb[self.TRANSACTIONS_SHEET]
        
        current_year = datetime.now().strftime("%Y")
        from collections import defaultdict
        cat_expense = defaultdict(float)
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]:
                continue
            if str(row[0])[:4] == current_year:
                amount = float(row[4] or 0)
                if amount < 0:
                    cat_expense[row[1] or "อื่นๆ"] += abs(amount)
        
        if not cat_expense:
            return f"📆 ยังไม่มีข้อมูลปี {current_year}"
        
        total = sum(cat_expense.values())
        lines = [f"📆 *สรุปปี {current_year}*\n", "💸 *รายจ่ายแยกตามหมวด:*"]
        
        for cat, amt in sorted(cat_expense.items(), key=lambda x: -x[1]):
            pct = amt / total * 100
            bar = "█" * int(pct / 5)
            lines.append(f"  {cat}\n  {bar} {amt:,.0f} บาท ({pct:.1f}%)")
        
        lines.append(f"\n💰 *รวมทั้งปี: {total:,.0f} บาท*")
        lines.append(f"📉 *เฉลี่ยต่อเดือน: {total/12:,.0f} บาท*")
        
        return "\n".join(lines)

    def get_recent_transactions(self, n: int = 10) -> list:
        """คืนค่า n รายการล่าสุด พร้อม row number"""
        wb = load_workbook(self.filepath, data_only=True)
        ws = wb[self.TRANSACTIONS_SHEET]
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=False):
            if row[0].value:
                rows.append((
                    row[0].row,
                    str(row[0].value)[:10],
                    row[1].value or "",
                    row[2].value or "",
                    row[3].value or "",
                    float(row[4].value or 0)
                ))
        return rows[-n:][::-1]  # ล่าสุดก่อน

    def delete_transaction(self, row_num: int):
        """ลบแถวที่ระบุแล้วอัพเดท summary"""
        wb = load_workbook(self.filepath)
        ws = wb[self.TRANSACTIONS_SHEET]
        ws.delete_rows(row_num)
        wb.save(self.filepath)
        self._update_summary_sheets(wb)

    def update_amount(self, row_num: int, new_amount: float):
        """แก้ไขจำนวนเงินในแถวที่ระบุ"""
        wb = load_workbook(self.filepath)
        ws = wb[self.TRANSACTIONS_SHEET]
        cell = ws.cell(row=row_num, column=5)
        # เก็บ sign เดิมไว้ (รายจ่ายเป็นลบ รายได้เป็นบวก)
        old_val = float(cell.value or 0)
        sign = -1 if old_val < 0 else 1
        cell.value = sign * new_amount
        wb.save(self.filepath)
        self._update_summary_sheets(wb)

    def update_date(self, row_num: int, new_date: str):
        """แก้ไขวันที่ในแถวที่ระบุ"""
        wb = load_workbook(self.filepath)
        ws = wb[self.TRANSACTIONS_SHEET]
        ws.cell(row=row_num, column=1).value = new_date
        wb.save(self.filepath)
        self._update_summary_sheets(wb)
